import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import openpyxl

# Define your original functions here, such as process_columns, save_avg_to_another_cell, etc.

# Example function definitions
def process_columns():
    # Insert the original code for process_columns here
    sheet2_row = 2  # Start writing results to Sheet2 from this row
    for col in range(1, 23):
        if col in {1, 2, 3, 4, 5, 6, 11, 12, 13, 14, 15, 16}:
            threshold = 0.6 * 2
        elif col in {7, 8, 9, 10, 17, 18, 19, 20}:
            threshold = 0.6 * 5
        elif col in {21}:
            threshold = 0.6 * 80
        elif col in {22}:
            threshold = 0.6 * 25
        else:
            continue

        count, attainment = calculate_attainment_level(col, threshold)
        save_attainment_results(col, count, attainment, sheet2_row)
        
        empty_count = count_empty_or_null_cells(col)
        sheet2.cell(row=sheet2_row + 3, column=col + 2, value=empty_count)
def count_empty_or_null_cells(col):
    empty_count = 0
    for cell in sheet1.iter_cols(min_col=col + 2, max_col=col + 2, min_row=1 + 3, max_row=sheet1.max_row):
        for col_cell in cell:
            if col_cell.value is None or col_cell.value == 0:
                empty_count += 1
    return empty_count
def calculate_attainment_level(col, threshold):
    count = 0
    for cell in sheet1.iter_cols(min_col=col + 2, max_col=col + 2, min_row=1 + 3, max_row=sheet1.max_row):
        for col_cell in cell:
            if isinstance(col_cell.value, (int, float)):  # Ensure the cell contains a number
                if col_cell.value >= threshold:  # Compare the cell value with the threshold
                    count += 1
    # Calculate attainment level
    attainment = int((count / No_of_Students) * 100)
    return count, attainment
def save_attainment_results(col, count, attainment, sheet2_row):
    # Save the data as numeric types
    sheet2.cell(row=sheet2_row, column=col + 2, value=count)
    sheet2.cell(row=sheet2_row + 1, column=col + 2, value=attainment)
    # Save the attainment level
    if attainment_11 > attainment >= attainment_01:
        sheet2.cell(row=sheet2_row + 2, column=col + 2, value=1)
    elif attainment_02 > attainment >= attainment_11:
        sheet2.cell(row=sheet2_row + 2, column=col + 2, value=2)
    elif attainment >= attainment_02:
        sheet2.cell(row=sheet2_row + 2, column=col + 2, value=3)
    else:
        sheet2.cell(row=sheet2_row + 2, column=col + 2, value=0)
def calculate_avg_of_columns_in_row(sheet2, columns, row):
    total = 0
    count = 0
    for col in columns:
        cell_value = sheet2.cell(row=row, column=col).value
        if isinstance(cell_value, (int, float)):  # Ensure it's a number
            total += cell_value
            count += 1  # Only count valid numbers
    if count == 0:
        return 0  # Avoid division by zero
    avg = total / count
    return avg
def save_avg_to_another_cell():
    row_to_calculate = 4  # Example row

    # Define column groups
    Co1 = [3, 4, 5, 6, 11, 12]  # Columns for first average
    Co2 = [7, 8, 9, 10]         # Columns for second average
    Co3 = [17, 22]              # Columns for third average
    Co4 = [14, 16, 20]          # Columns for fourth average
    Co5 = [15, 18, 19, 21]      # Columns for fifth average
    Co6 = [13]                  # Columns for sixth average

    # Calculate averages for each column group
    avg1 = calculate_avg_of_columns_in_row(sheet2, Co1, row_to_calculate)
    avg2 = calculate_avg_of_columns_in_row(sheet2, Co2, row_to_calculate)
    avg3 = calculate_avg_of_columns_in_row(sheet2, Co3, row_to_calculate)
    avg4 = calculate_avg_of_columns_in_row(sheet2, Co4, row_to_calculate)
    avg5 = calculate_avg_of_columns_in_row(sheet2, Co5, row_to_calculate)
    avg6 = calculate_avg_of_columns_in_row(sheet2, Co6, row_to_calculate)

    # Round the averages to 2 decimal places
    avg1_float = round(avg1, 2)
    avg2_float = round(avg2, 2)
    avg3_float = round(avg3, 2)
    avg4_float = round(avg4, 2)
    avg5_float = round(avg5, 2)
    avg6_float = round(avg6, 2)

    # Save the averages to specific cells in Sheet2
    sheet2.cell(row=12, column=3, value=avg1_float)
    sheet2.cell(row=13, column=3, value=avg2_float)
    sheet2.cell(row=14, column=3, value=avg3_float)
    sheet2.cell(row=15, column=3, value=avg4_float)
    sheet2.cell(row=16, column=3, value=avg5_float)
    sheet2.cell(row=17, column=3, value=avg6_float)
def calculate_external_avg(sheet2, column_value, row):
    #(0.3(C12+X3)/2 + 0.7(W3))
    X3 = sheet2.cell(row=row, column=24).value
    W3 = sheet2.cell(row=row, column=23).value
    
    # Ensure X3, W3, and column_value are numbers before performing calculations
    if isinstance(X3, (int, float)) and isinstance(W3, (int, float)) and isinstance(column_value, (int, float)):
        Avg_1 = (0.3 * (column_value + X3) / 2) + (0.7 * W3)
        return Avg_1
    else:
        print(f"Invalid values in row {row}. X3: {X3}, W3: {W3}, Column Value: {column_value}")
        return None
def save_external_avg():
    # Insert the original code for save_external_avg here
    row_to_calculate = 4  # The row to fetch X3 and W3 values

    # Fetch previously saved averages from Sheet2
    Co1 = sheet2.cell(row=12, column=3).value
    Co2 = sheet2.cell(row=13, column=3).value
    Co3 = sheet2.cell(row=14, column=3).value
    Co4 = sheet2.cell(row=15, column=3).value
    Co5 = sheet2.cell(row=16, column=3).value
    Co6 = sheet2.cell(row=17, column=3).value

    # Calculate external averages for each column
    cal1 = calculate_external_avg(sheet2, Co1, row_to_calculate)
    cal2 = calculate_external_avg(sheet2, Co2, row_to_calculate)
    cal3 = calculate_external_avg(sheet2, Co3, row_to_calculate)
    cal4 = calculate_external_avg(sheet2, Co4, row_to_calculate)
    cal5 = calculate_external_avg(sheet2, Co5, row_to_calculate)
    cal6 = calculate_external_avg(sheet2, Co6, row_to_calculate)

    # Round the calculated values to 2 decimal places if they are valid
    if cal1 is not None:
        cal1_float = round(cal1, 2)
        sheet2.cell(row=12, column=4, value=cal1_float)  # Save in a different column
    if cal2 is not None:
        cal2_float = round(cal2, 2)
        sheet2.cell(row=13, column=4, value=cal2_float)
    if cal3 is not None:
        cal3_float = round(cal3, 2)
        sheet2.cell(row=14, column=4, value=cal3_float)
    if cal4 is not None:
        cal4_float = round(cal4, 2)
        sheet2.cell(row=15, column=4, value=cal4_float)
    if cal5 is not None:
        cal5_float = round(cal5, 2)
        sheet2.cell(row=16, column=4, value=cal5_float)
    if cal6 is not None:
        cal6_float = round(cal6, 2)
        sheet2.cell(row=17, column=4, value=cal6_float)

def calculate_avg_and_save(sheet2, start_col, end_col, start_row, end_row, avg_row):
    # Insert the original code for calculate_avg_and_save here
    for col in range(start_col, end_col + 1):
        total = 0
        count = 0
        
        # Loop through each row from start_row to end_row (13 to 18)
        for row in range(start_row, end_row + 1):
            cell_value = sheet2.cell(row=row, column=col).value
            if isinstance(cell_value, (int, float)):  # Only count numeric values
                total += cell_value
                count += 1
        
        # Calculate the average and print it
        if count > 0:
            avg = total / count
            avg_round = round(avg,2)
        else:
            avg_round = 0  # Avoid division by zero in case there are no numeric values
        sheet2.cell(row=avg_row, column=col, value=avg_round)        
def co_table1():
    new_table_start_row = 12
    new_table_start_col = 7

# Variables for the source table (used for 'val' fetching)
    cols = 4  # Start column for 'val'
    mcols = 3  # End column for 'val'
    rows = 12  # Start row for 'val'
    mrows = 17  # End row for 'val'

# Iterate through the range (columns 6 to 20, rows 13 to 19)
    for col in range(28, 42):
        for row in range(5, 11):
            cell = sheet1.cell(row=row, column=col)  # Current cell in the original table
            cell_value = cell.value  # Get the current cell value

        # Fetch 'val' from the static range (cols, mcols, rows, mrows)
            val = sheet2.cell(row=row + rows - 5, column=cols).value  # Adjusting row to match rows variable

        # Apply conditional logic
            if isinstance(cell_value, (int, float)):
                if cell_value == 3:
                    new_value = val  # Use 'val' as is
                elif cell_value == 2:
                    new_value = (val)* 0.66  # Scale 'val'
                elif cell_value == 1:
                    new_value = (val)* 0.33  # Scale 'val'
                else:
                    new_value = 0  # Default for other numbers
            else:
                new_value = ' '  # If the original cell is not numeric

        # Write the new value in the new table (Sheet2)
            new_row = new_table_start_row + (row - 5)  # Adjust new row index for the new table
            new_col = new_table_start_col + (col - 28)  # Adjust new column index for the new table
            sheet2.cell(row=new_row, column=new_col, value=new_value)  # Save the transformed value

def co_table2():
    rows = 12
    cols = 4  # Adjust based on where 'val' is coming from
    new_table_start_row = 12
    new_table_start_col = 23

    # Iterate through the range (columns 7 to 20, rows 12 to 17)
    for col in range(28, 42):
        for row in range(5, 11):
            # Get the cell from the current table (Sheet2)
            cell = sheet1.cell(row=row, column=col)
            cell_value = cell.value  # Extract the value of the current cell

            # Extract the value of 'val' from the source (rows and cols)
            val = sheet2.cell(row=row + rows - 5, column=cols).value  # Get the value of 'val'

            # Apply the conditions
            if isinstance(cell_value, (int, float)):  # Ensure it's a number
                if cell_value == 3 or cell_value == 1 or cell_value == 2:
                    new_value = val  # Use 'val' if condition is met
                else:
                    new_value = ' '  # Empty space for other conditions
            else:
                new_value = ' '  # Non-numeric values default to empty space
            # Calculate new row and column for the new table
            new_row = new_table_start_row + (row - 5)  # Adjust for the new table's starting row
            new_col = new_table_start_col + (col - 28)  # Adjust for the new table's starting column

            # Write the new value to the new table
            sheet2.cell(row=new_row, column=new_col, value=new_value)  # Save the transformed value
# Now, integrate the UI code as before
filepath = None
workbook = None
sheet1 = None
sheet2 = None
def load_workbook_file():
    global filepath, workbook, sheet1, sheet2
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if filepath:
        try:
            workbook = load_workbook(filepath)
            sheet1 = workbook['Sheet1']
            sheet2 = workbook['Sheet2']
            last_row = sheet1.max_row
            messagebox.showinfo("Success", f"File {filepath} loaded successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Error loading workbook: {e}")
    else:
        messagebox.showwarning("Warning", "No file selected.")
def calculate_students():
    global No_of_Students
    if sheet1:
        No_of_Students = sheet1.max_row - 3
        student_label.config(text=f"Total Students: {No_of_Students}")
    else:
        messagebox.showwarning("Warning", "No file loaded.")
def set_attainment_values():
    global attainment_01, attainment_11, attainment_02
    try:
        attainment_01 = int(attain_min_entry.get())
        attainment_11 = int(attain_max_entry.get())
        attainment_02 = int(attain_second_max_entry.get())
        messagebox.showinfo("Success", "Attainment values set successfully.")
    except ValueError:
        messagebox.showerror("Error", "Please enter valid integers.")
def run_processing():
    if sheet1 and sheet2:
        try:
            process_columns()
            save_avg_to_another_cell()
            save_external_avg()
            calculate_avg_and_save(sheet1, start_col=28, end_col=41, start_row=5, end_row=10, avg_row=12)
            co_table1()
            calculate_avg_and_save(sheet2, start_col=7, end_col=20, start_row=12, end_row=17, avg_row=19)
            co_table2()
            calculate_avg_and_save(sheet2, start_col=23, end_col=36, start_row=12, end_row=17, avg_row=19)
            workbook.save(filepath)
            workbook.close()
            messagebox.showinfo("Success", "Data processed and saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during processing: {e}")
    else:
        messagebox.showwarning("Warning", "No file loaded or attainment values not set.")
# Tkinter GUI Setup
root = tk.Tk()
root.title("CO-PO Attainment Evaluation")
root.geometry("400x400")
load_button = tk.Button(root, text="Load Excel File", command=load_workbook_file)
load_button.pack(pady=10)
calculate_button = tk.Button(root, text="Calculate Students", command=calculate_students)
calculate_button.pack(pady=10)
student_label = tk.Label(root, text="Total Students: ")
student_label.pack()
attain_label = tk.Label(root, text="Enter Attainment Values:")
attain_label.pack(pady=5)
attain_min_label = tk.Label(root, text="Min Range:")
attain_min_label.pack()
attain_min_entry = tk.Entry(root)
attain_min_entry.pack()
attain_max_label = tk.Label(root, text="Max Range for Attainment 1:")
attain_max_label.pack()
attain_max_entry = tk.Entry(root)
attain_max_entry.pack()
attain_second_max_label = tk.Label(root, text="Max Range for Attainment 2:")
attain_second_max_label.pack()
attain_second_max_entry = tk.Entry(root)
attain_second_max_entry.pack()
set_button = tk.Button(root, text="Set Attainment Values", command=set_attainment_values)
set_button.pack(pady=10)
process_button = tk.Button(root, text="Process Data", command=run_processing)
process_button.pack(pady=10)
exit_button = tk.Button(root, text="Exit", command=root.quit)
exit_button.pack(pady=10)
root.mainloop()